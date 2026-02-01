"""Utility helpers for NEHVI experiments.

This module provides small utilities used by the NEHVI workflows and the corresponding BO (Bayesian Optimisation) experiments:
- populating Excel input templates for the DT simulator
- evaluating the DT model via an external JAR
- reading KPI metrics from the simulation outputs
- orchestrating batches of simulations
"""

from openpyxl import load_workbook
import subprocess
import pprint
import csv
from multiprocessing import cpu_count, Pool
from functools import partial

pp = pprint.PrettyPrinter(width=41, compact=True)


def write_to_input_file(excel_file, testing=False, **kwargs):
    """Populate the Excel input template with given design values.

    This helper writes values into predefined cells in the workbook for the
    simulation run. Keys in kwargs must match the expected input names used in
    the template (e.g. chute_east_zone_cluster_res, chute_west_zone_cluster_res,
    infeed_south_zone_cluster_res, infeed_north_zone_cluster_res, max_rc).

    Args:
      excel_file (str): Path to the input Excel workbook to modify.
      testing (bool): If True, configure the workbook for a short test run.
      **kwargs: Design input values used to populate the template.

    Returns:
      None
    """
    print("testing = ", testing)

    try:
        wb = load_workbook(excel_file)
    except EOFError:
        print(excel_file)

    chute_res_sheet_name = "ChuteResource"
    infeed_res_sheet_name = "InfeedResource"
    inputs_sheet_name = "Inputs"

    assert chute_res_sheet_name in wb.sheetnames
    assert infeed_res_sheet_name in wb.sheetnames
    assert inputs_sheet_name in wb.sheetnames

    chute_sheet = wb[chute_res_sheet_name]
    chute_sheet["C3"].value = kwargs["chute_east_zone_cluster_res"]
    chute_sheet["C5"].value = kwargs["chute_west_zone_cluster_res"]

    infeed_sheet = wb[infeed_res_sheet_name]
    infeed_sheet["C2"].value = kwargs["infeed_south_zone_cluster_res"]
    infeed_sheet["C3"].value = kwargs["infeed_north_zone_cluster_res"]

    wb[inputs_sheet_name]["B17"].value = kwargs["max_rc"]

    # wb["Config"]["B7"].value = "SyntheticCreation"
    # wb["Config"]["B8"].value = 5
    wb["Resource"]["B5"].value = 0.5  # productivity multiplier for chute resources
    wb["Resource"]["B11"].value = 0.6  # productivity multiplier for infeed resources

    wb["Config"]["B5"].value = 600

    if testing:
        wb["Config"]["B5"].value = 5  # run for only 5 ticks

    wb.save(excel_file)
    print("Finished writing new inputs to file.")



from pathlib import Path

def dt_evaluation(io_file, args, results_dir):
    if args.env == "welkin":
        rootdir = Path("/data/ML/Garima/Digital-twin/IO/Koge")
        jar_filepath = Path("/data/ML/Garima/Digital-twin/jars/May20.jar")
    else:
        # Use the correct Windows path; Path handles backslashes or slashes
        rootdir = Path(r"C:\Users\2845374\OneDrive - TCS COM PROD\Desktop\dt-optimization\IO\Koge")
        jar_filepath = Path(r"C:\Users\2845374\OneDrive - TCS COM PROD\Desktop\dt-optimization\model.jar")

    parcel_load_sheet = rootdir / "s1-Parcel_data.csv"
    parcel_chute_mapping_sheet = rootdir / "parcel_chute_cage_mapping.xlsx"

    cmd = [
        "java",
        "-jar",
        str(jar_filepath),
        "false",
        str(io_file[0]),
        str(io_file[1]),
        "Standard",
        str(parcel_load_sheet),
        str(parcel_chute_mapping_sheet),
    ]

    attempt = 1
    max_tries = 5
    run_complete = False

    while attempt <= max_tries:
        try:
            # Ensure results_dir exists
            results_dir = Path(results_dir)
            results_dir.mkdir(parents=True, exist_ok=True)
            with open(results_dir / "dt_out.txt", "+a") as file:
                timeout = 30 if args.testing else 300
                result = subprocess.run(cmd, timeout=timeout, stdout=file, stderr=file)
            run_complete = True
            if result.returncode != 0:
                print(f"Return code = {result.returncode}")
                raise SystemExit(1)
            break
        except subprocess.TimeoutExpired:
            attempt += 1
            print(f"Attempt {attempt}")

    if not run_complete:
        print("Max attempts exceeded. Run the job again.")
        raise SystemExit(1)

    print("Run complete.")

def read_kpi_from_out_file(excel_file: str, total_resources: int):
    """Read KPI metrics from a DT output workbook.

    This function extracts key performance indicators such as throughput,
    dwelling time and recirculation counts from the Excel sheets produced by
    the simulator.

    Returns:
      A tuple:
        (scanned_parcels, sorted_parcels, rejected_parcels,
         throughput, avg_dwelling_time, avg_recirculations, reject_rate)
    """
    """
    KPIs needed:
    - throughput -> no. of sorted parcels / total incoming parcels
    - avg. no of recirculations -> total no. of recirculations / total scanned parcels(sorted+rejected)
    - rate of rejections -> no. of rejected parcels / total scanned parcels(sorted+rejected)
    - avg. dwelling time -> avg. time spent by parcel(journey from conveyor belt to chute/cage)
    """
    wb = load_workbook(excel_file)

    sheet1 = "lastTick"
    sheet2 = "cumulativeData"
    sheet3 = "Recirculation Count Data"
    sheet4 = "Chutewise parcels pushed"

    assert sheet1 in wb.sheetnames
    assert sheet2 in wb.sheetnames
    assert sheet3 in wb.sheetnames
    assert sheet4 in wb.sheetnames

    scanned_parcels = int(wb[sheet2][2][-1].value)  # read the last value of 2nd row in sheet2

    recirculation_counts = [int(cell.value) for cell in wb[sheet3]["A"][1:]]  # col A without heading
    parcels_recirculated_counts = [int(cell.value) for cell in wb[sheet3]["B"][1:]]  # col B without heading
    parcels_recirculated_counts_sum = sum(
        parcels_recirculated_counts
    )  # sorted + rejected + still left at chutes awaiting cage allocation

    assert len(recirculation_counts) == len(parcels_recirculated_counts)

    # last value in Sheet 3, 2nd col
    rejected_parcels = int(wb[sheet3].cell(row=len(recirculation_counts) + 1, column=2).value)

    # cell value is the total parcels, which either got sorted or rejected
    sorted_parcels = int(wb[sheet1]["D4"].value) - rejected_parcels  #  parcels alloted to cage for outbound

    avg_recirculations = round(
        (
            sum(recirculation_counts[i] * parcels_recirculated_counts[i] for i in range(len(recirculation_counts)))
            / parcels_recirculated_counts_sum
        ),
        4,
    )

    throughput = round(sorted_parcels / 57383, 4)

    reject_rate = round(rejected_parcels / (parcels_recirculated_counts_sum), 4)

    # avg dwelling time = total dwelling time / no. of parcels(assigned to cages + rejected)
    avg_dwelling_time = round(wb[sheet1]["B4"].value, 4)

    print(
        "#########################################################################################################################################\n"
        f"Sorted parcels = {sorted_parcels} \t Rejected parcels = {rejected_parcels}"
        "                                                                   KPIs OBTAINED \n"
        f" 1. THROUGHPUT = {throughput}\n"
        f" 2. TOTAL RESOURCES = {total_resources}\n"
        f" 3. AVG. RECIRCULATIONS = {avg_recirculations} \n"
        f" 4. REJECT RATE = {reject_rate}\n"
        f" 5. AVG. DWELLING TIME = {avg_dwelling_time}\n\n"
    )

    return (
        scanned_parcels,
        sorted_parcels,
        rejected_parcels,
        throughput,
        avg_dwelling_time,
        avg_recirculations,
        reject_rate,
    )


def run_simulation(
    args,
    results_dir,
    output_dir,
    batch_size,
    DT_SIMULATIONS,
    total_res_list,
    inputs,
    exp2=False,
):
    """Run DT simulation given new inputs, and return objective functions."""
    input_files = [args.input_files[0]] * batch_size
    for i in range(batch_size):
        write_to_input_file(input_files[i], testing=True, **inputs[i])

    # run system command to run the DT simulation with new inputs
    print("starting a simulation..")

    obj_fn_values = []

    with open(f"{results_dir}/simulations.csv", "+a") as file:
        writer = csv.writer(file, delimiter=" ")

        print("No. of cpus = ", cpu_count())
        output_files = [f"{output_dir}/out{i + 1}.xlsx" for i in range(batch_size)]

        io_files = [[input_file, output_file] for (input_file, output_file) in zip(input_files, output_files)]

        with Pool(batch_size) as pool:
            partial_func = partial(dt_evaluation, args=args, results_dir=results_dir)
            _ = pool.map(partial_func, io_files)

        for i in range(batch_size):
            DT_SIMULATIONS += 1
            (
                scanned_parcels,
                sorted_parcels,
                rejected_parcels,
                throughput,
                dwelling_time,
                recirculations,
                reject_rate,
            ) = read_kpi_from_out_file(
                excel_file=output_files[i],
                total_resources=total_res_list[i],
            )
            if exp2:  # don't add resources to objectives
                obj_fn_values.append([throughput, -dwelling_time, -recirculations, -reject_rate])
            else:
                obj_fn_values.append([throughput, -dwelling_time, -recirculations, -reject_rate, -total_res_list[i]])

            writer.writerow(
                [
                    DT_SIMULATIONS,
                    total_res_list[i],
                    inputs[i]["chute_east_zone_cluster_res"],
                    inputs[i]["chute_west_zone_cluster_res"],
                    inputs[i]["infeed_north_zone_cluster_res"],
                    inputs[i]["infeed_south_zone_cluster_res"],
                    inputs[i]["max_rc"],
                    throughput,
                    reject_rate,
                    dwelling_time,
                    recirculations,
                    scanned_parcels,
                    sorted_parcels,
                    rejected_parcels,
                ]
            )

    return obj_fn_values
